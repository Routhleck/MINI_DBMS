import os
import sys
#导入QFileDialog
from PyQt5.QtWidgets import QFileDialog

import qdarkstyle
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtGui import QTextCursor
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QTableWidgetItem
from openpyxl import load_workbook

import UI_lan
import main_interpreter
from utils import assign_style_qt, get_merge_cell_list

username = ''
password = ''
flagFirst = True
flagLogin = False
filePath = 'data/'


class Stream(QObject):
    """Redirects console output to text widget."""
    newText = pyqtSignal(str)

    def write(self, text):
        QtWidgets.QApplication.processEvents()
        self.newText.emit(str(text))


class anaxcelhandler(QtWidgets.QMainWindow, UI_lan.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        # self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        # self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        # self.setStyleSheet(open("qss/1.qss", "r").read())

        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonbrowseLoad.clicked.connect(self.openFileNamesDialogLoad)
        self.pushButtonclear.clicked.connect(self.clearwidget)
        self.pushButtonselall.clicked.connect(self.selectall)
        self.pushButtonload.clicked.connect(self.LoadProcess)
        self.pushButton_submit.clicked.connect(self.submit)

        self.statusbar.showMessage('Mini DBMS by group 10')
        self.comboBoxfiletype.addItems(['xlsx', 'xls'])

        # ==========log=====
        sys.stdout = Stream(newText=self.onUpdateText)
        # ==========log=====

        # ==========show====
        self.flag_confirm = False
        self.activate_file = [None, None]
        self.comboBox_wb.activated.connect(self.wbActivated)
        self.comboBox_ws.activated.connect(self.wsActivated)
        self.tableWidget.itemClicked.connect(self.handleItemClick)
        self.pushButton_clear_idx.clicked.connect(self.clear_idx)
        self.pushButton_confirm_idx.clicked.connect(self.confirm_idx)
        # ==========show====

        # ==========context===
        self.infos = {}
        self.infos_bak = {}

    def submit(self):
        # 将lineEdit_input的内容传入函数中
        global username
        global password
        global flagFirst
        global flagLogin
        if flagFirst:
            print('请先登录')
            print('用户名:')
            flagFirst = False
        elif username == '':
            if self.lineEdit_input.text() == '':
                print('\n用户名为空，请重新输入:')
            else:
                username = self.lineEdit_input.text()
                print(username)
                print('密码:')
        elif not username == '' and password == '':
            if self.lineEdit_input.text() == '':
                print('\n密码为空，请重新输入:')
            else:
                password = self.lineEdit_input.text()
                print(password)
                print('再次点击进行验证')
        elif not username == '' and not password == '' and not flagLogin:
            flagFirst, flagLogin = main_interpreter.userLogin(username, password, flagFirst, flagLogin)
            if not flagLogin:
                username = ''
                password = ''
        elif flagLogin:
            main_interpreter.interpreter(self.lineEdit_input.text())

    def use_palette(self):
        self.setWindowTitle("设置背景图片")
        window_pale = QtGui.QPalette()
        window_pale.setBrush(self.backgroundRole(), QtGui.QBrush(QtGui.QPixmap(self.bundle_dir + '/icons/bg.jpeg')))
        self.setPalette(window_pale)

    def onUpdateText(self, text):
        """Write console output to text widget."""
        cursor = self.textBrowserlog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowserlog.setTextCursor(cursor)
        self.textBrowserlog.ensureCursorVisible()

    def openFileNamesDialog(self):
        # 获取filePath所有xlsx文件存储到files中
        global filePath
        files = []
        for file in os.listdir(filePath):
            if file.endswith('.xlsx'):
                files.append(file)

        """options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    #仅保留file的文件名且去除后缀
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    #仅保留file的文件名且去除后缀
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        """
        if files:
            for file in files:
                # 仅保留file的文件名且去除后缀
                self.listWidget.addItem(os.path.basename(file).split('.')[0])

    def openFileNamesDialogLoad(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_all(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_show(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)

    def clear_idx(self):
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def assign_dict(self, dict1, dict2):
        for k, v in dict1.items():
            if isinstance(v, dict):
                dict_tmp = dict()
                dict2[k] = self.assign_dict(v, dict_tmp)
            else:
                dict2[k] = v
        return dict2

    def confirm_idx(self):
        self.infos_bak = self.assign_dict(self.infos, self.infos_bak)

        x = self.comboBox_x.itemText(self.comboBox_x.currentIndex())
        y = self.comboBox_y.itemText(self.comboBox_y.currentIndex())

        r1 = self.comboBox_r1.itemText(self.comboBox_r1.currentIndex())
        r2 = self.comboBox_r2.itemText(self.comboBox_r2.currentIndex())

        wb = self.comboBox_wb.itemText(self.comboBox_wb.currentIndex())
        ws = self.comboBox_ws.itemText(self.comboBox_ws.currentIndex())

        if wb == '' or ws == '':
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '先load文件')
        else:
            x = int(x) if x != '' else x
            y = int(y) if y != '' else y
            r1 = int(r1) if r1 != '' else r1
            r2 = int(r2) if r2 != '' else r2

            if self.checkBox_book.isChecked():
                print('book')
                key_idx = [x, y]
                rg = [r1, 'last']
                for wb_k in self.infos_bak.keys():
                    ws_keys = self.infos_bak[wb_k]['sheet_names']
                    for ws_k in ws_keys.keys():
                        self.infos_bak[wb_k]['sheet_names'][ws_k] = [key_idx, rg]
            elif self.checkBox_sheet.isChecked():
                print('sheet')
                key_idx = [x, y]
                rg = [r1, 'last']
                ws_keys = self.infos_bak[wb]['sheet_names']
                for ws_k in ws_keys.keys():
                    self.infos_bak[wb]['sheet_names'][ws_k] = [key_idx, rg]
            else:
                print('cell')
                key_idx = [x, y]
                rg = [r1, r2]
                self.infos_bak[wb]['sheet_names'][ws] = [key_idx, rg]
            self.flag_confirm = True

    def selectall(self):
        self.listWidget.selectAll()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '请先加载文件')

    def LoadProcess(self):
        self.clearcontext_all()
        if self.comboBoxfiletype.currentIndex() == 1:  # xls
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '不支持 xls 格式文件')
        elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
            items = self.listWidget.selectedItems()
            if len(items) == 0:
                QMessageBox.about(self, "hi,Mini DBMS by group 10", '请先选择文件')
            else:
                self.infos = {}
                for i in list(items):
                    file_path = str(filePath + i.text() + '.xlsx')
                    wb = load_workbook(filename=file_path)
                    name = os.path.split(file_path)[-1]

                    sheet_names = wb.sheetnames

                    sheets_dict = {}
                    for s in sheet_names:
                        sheets_dict[s] = []
                    self.infos[name] = {'path': file_path, 'sheet_names': sheets_dict}
                    wb.close()
                for k in self.infos.keys():
                    self.comboBox_wb.addItem(k)
                k = self.comboBox_wb.itemText(0)
                sheets = list(self.infos[k]['sheet_names'].keys())
                for s in sheets:
                    self.comboBox_ws.addItem(s)
                self.activate_file[0] = self.infos[k]['path']
                self.activate_file[1] = list(self.infos[k]['sheet_names'].keys())[0]

                self.show_excel()
        print('可以预览文件')

    def wbActivated(self, index):
        self.clearcontext_show()
        wb_k = self.comboBox_wb.itemText(index)
        sheets = list(self.infos[wb_k]['sheet_names'].keys())
        self.comboBox_ws.clear()
        for s in sheets:
            self.comboBox_ws.addItem(s)
        self.activate_file[0] = self.infos[wb_k]['path']
        self.activate_file[1] = list(self.infos[wb_k]['sheet_names'].keys())[0]
        self.show_excel()

    def wsActivated(self, index):
        ws_k = self.comboBox_ws.itemText(index)
        self.activate_file[1] = ws_k
        self.show_excel()

    def handleItemClick(self, item):
        cont = item.text()
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        row = item.row() + 1
        column = item.column() + 1
        # =======对合并的单元格取idx
        for p in self.merge_position:
            if row == p[0] and column == p[1]:
                row = row + (p[2] - p[0])
                break
        # =======对合并的单元格取idx
        self.comboBox_x.addItem(str(row))
        self.comboBox_y.addItem(str(column))
        self.comboBox_r1.addItem(str(row + 1))

    def show_excel(self):
        self.merge_position = []
        path = self.activate_file[0]
        sheetname = self.activate_file[1]
        wb = load_workbook(filename=path)
        ws = wb[sheetname]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)

        # ======合并单元格=======
        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0] - 1, m_idx[1] - 1, m_idx[2] - m_idx[0] + 1, m_idx[3] - m_idx[1] + 1)
            self.merge_position.append([m_idx[0], m_idx[1], m_idx[2]])  # [x1,y1,range]
        # ======合并单元格=======

        # ======单元格大小=======
        for i in range(1, num_row + 1):
            h = ws.row_dimensions[i].height
            if h is not None:
                self.tableWidget.setRowHeight(i - 1, h)
        # for i in range(1,num_column+1):
        #     w = ws.column_dimensions[get_column_letter(i)].width
        #     if w is not None:
        #         self.tableWidget.setColumnWidth(i-1,w)
        # ======单元格大小=======

        self.comboBox_r2.clear()
        for i in range(1, num_row + 1):
            self.comboBox_r2.addItem(str(num_row - i + 1))
            row_sizes = []
            for j in range(1, num_column + 1):
                cell = ws.cell(row=i, column=j)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))
                    assign_style_qt(item, cell)
                else:
                    item = QTableWidgetItem()
                self.tableWidget.setItem(i - 1, j - 1, item)

        # self.tableWidget.resizeColumnsToContents()
        # self.tableWidget.resizeRowsToContents()
        wb.close()


app = QtWidgets.QApplication(sys.argv)
window = anaxcelhandler()
window.openFileNamesDialog()
# setup stylesheet
app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
window.show()
sys.exit(app.exec_())
