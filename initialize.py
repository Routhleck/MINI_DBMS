import hashlib
import os
import re
from openpyxl import *
import dbms_function

db_path = 'data/'
# view_path = 'view/'
user = ''
using_dbname = ''
using_db = Workbook()


# 创建数据库
def creat_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = Workbook()
    database.save(dbpath)

    #为initialize首次创建system[permission]表
    #system.xlsx中创建permission sheet
    permission = database.create_sheet('permission')
    permission.cell(row=1, column=1).value = 'database'
    permission.cell(row=1, column=2).value = 'select'
    permission.cell(row=1, column=3).value = 'insert'
    permission.cell(row=1, column=4).value = 'delete'
    permission.cell(row=1, column=5).value = 'update'
    permission.cell(row=1, column=6).value = 'use'
    #保存
    database.save(dbpath)

    #system.xlsx中创建user sheet
    user = database.create_sheet('user')
    user.cell(row=1, column=1).value = 'username'
    user.cell(row=1, column=2).value = 'password'
    user.cell(row=2, column=1).value = 'admin'
    user.cell(row=2, column=2).value = hashlib.md5('admin'.encode('utf-8')).hexdigest()
    #保存
    database.save(dbpath)

    dbms_function.create_tb_in_tbinfo(dbname)
    print(u"数据库创建操作执行成功")

def Initialization():
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists("data/table_information.xlsx"):
        Workbook().save("data/table_information.xlsx")
    if os.path.exists("data/system.xlsx"):
        print("Initializating......")
    else:
        creat_db('system')
    db = load_workbook("data/system.xlsx")
    permission_tb_col = ['database char[50] pk unique', 'select char', 'insert char', 'delete char', 'update char']
    dbms_function.creat_table('permission', db, 'system', permission_tb_col)

#主函数
if __name__ == '__main__':
    Initialization()
    #run()