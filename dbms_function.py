from openpyxl import *
import  os
import re
from index import *
from prettytable import PrettyTable
import hashlib
db_path = 'data/'


#在table_infomation中创建数据库对应的表
def create_tb_in_tbinfo(dbname):    
    db = load_workbook("data/table_information.xlsx")
    table = db.create_sheet(dbname)
    columns_name = ['table','column_name', 'type', 'null', 'unique', 'primary_key', 'foreign_key']
    for i in range(len(columns_name)):
        table.cell(row=1,column=i+1).value = columns_name[i]
    if db.worksheets[0].title == 'Sheet':
        del db['Sheet']
    db.save("data/table_information.xlsx")

    #创建数据库时，将数据库权限添加到system中
    db = load_workbook("data/system.xlsx")
    table = db['permission']
    #计算行数
    row_num = table.max_row + 1
    #从1开始至列数最大值
    for i in range(1,table.max_column+1):
        if i == 1:
            table.cell(row=row_num,column=i).value = dbname
        else:
            table.cell(row=row_num, column=i).value = 'admin,root'
    db.save("data/system.xlsx")


#create table tbname (id int PK null,user char[10] )
def creat_table(table_name,current_database,current_dbname,columns_list):
    # create table
    if table_name not in current_database.sheetnames:
        table = current_database.create_sheet(table_name)
    else:
        print (u"数据表已存在,请重新输入.")
        return
    if current_database.worksheets[0].title == 'Sheet':
        del current_database['Sheet']
    #表创建完成，开始创建列
    length = len(columns_list)
    #print length
    tbinfo = load_workbook("data/table_information.xlsx")
    tbinfo_tb = tbinfo[current_dbname]
    tbinfo_rows = tbinfo_tb.max_row
    column_names = []
    for i in range(length):             #将字段的属性写到table_information库中
        column = columns_list[i].split(' ')
        tbinfo_tb.cell(row=tbinfo_rows+1+i,column=1).value = table_name
        tbinfo_tb.cell(row=tbinfo_rows+1+i, column=2).value = column[0]
        tbinfo_tb.cell(row=tbinfo_rows+1+i, column=3).value = column[1]
        for key in column[2:]:
            if key == 'null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '1'
            elif key == 'not_null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '0'
            elif key == 'unique':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=5).value = '1'
            elif key == 'pk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=6).value = '1'
            elif key == 'fk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=7).value = '1'
        column_names.append(column[0])
        for j in range(1, 8):
            if tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value is None:
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value = 'NULL'
    tbinfo.save("data/table_information.xlsx")
    for i in range(length):
        table.cell(row=1,column=i+1).value = column_names[i]  #表第一行是列名
    current_dbname = db_path + current_dbname + '.xlsx'
    current_database.save(current_dbname)
    print (u"数据表创建完成。")

#删除数据库
def drop_db(dbname):
    #查看是否存在该数据库
    if os.path.exists(db_path+dbname+'.xlsx'):
        #删除文件
        os.remove(db_path+dbname+'.xlsx')
        print("数据库文件已删除.")
        #删除在table_information中的记录
        db = load_workbook("data/table_information.xlsx")
        worksheet = db[dbname]
        db.remove(worksheet)
        db.save("data/table_information.xlsx")
        print("数据库信息已删除.")

        #删除在system中的权限列
        db = load_workbook("data/system.xlsx")
        table = db['permission']
        

        #找到dbname的行
        for i in range(1,table.max_row+1):
            if table.cell(i,1).value == dbname:
                table.delete_rows(i,1)
                break
        print("数据库权限已删除.")

        db.save("data/system.xlsx")
    else:
        print("没有找到数据库.")
    
#删除表
def drop_table(tbname,using_dbname,using_db):
    #查看是否存在该表
    #打开table_information
    db = load_workbook("data/table_information.xlsx")
    #是否存在该表
    if using_dbname in db.sheetnames:
        worksheet = db[using_dbname]
        #删除所有第一列为tbname的行
        for i in range(1,worksheet.max_row+1):
            if worksheet.cell(i,1).value == tbname:
                worksheet.delete_rows(i,1)
        db.save("data/table_information.xlsx")
        print("table_information删除该表.")
    else:
        print("table_information中没有该数据库.")
    

    #打开using_dbname文件
    #该表是否在using_dbname文件中存在
    if os.path.exists(db_path+using_dbname+'.xlsx'):
        db = load_workbook(db_path+using_dbname+'.xlsx')
        #tbname是否存在于sheet中
        if tbname in db.sheetnames:
            #若db仅剩一个sheet，则删除该文件
            if len(db.sheetnames) == 1:
                os.remove(db_path+using_dbname+'.xlsx')
                print("因为该数据库仅剩一个表，所以将其删除.")
                return
            worksheet = db[tbname]
            db.remove(worksheet)
            db.save(db_path+using_dbname+'.xlsx')
            print("在数据库中删除该表.")
        else:
            print("该表不在数据库中.")
    else:
        print("未找到数据库.")
    
#添加字段
def add_field(tbname,columns_list,using_dbname,using_db):
    #查看是否存在该表
    #打开table_information
    db = load_workbook("data/table_information.xlsx")
    #是否存在该表
    if using_dbname in db.sheetnames:
        worksheet = db[using_dbname]

        length = len(columns_list)
        #添加新的一行
        tbinfo = db
        tbinfo_tb = tbinfo[using_dbname]
        tbinfo_rows = tbinfo_tb.max_row
        column_names = []
        for i in range(length):             #将字段的属性写到table_information库中
            column = columns_list[i].split(' ')
            tbinfo_tb.cell(row=tbinfo_rows+1+i,column=1).value = tbname
            tbinfo_tb.cell(row=tbinfo_rows+1+i, column=2).value = column[0]
            tbinfo_tb.cell(row=tbinfo_rows+1+i, column=3).value = column[1]
            for key in column[2:]:
                if key == 'null':
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '1'
                elif key == 'not_null':
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '0'
                elif key == 'unique':
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=5).value = '1'
                elif key == 'pk':
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=6).value = '1'
                elif key == 'fk':
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=7).value = '1'
            column_names.append(column[0])
            for j in range(1, 8):
                if tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value is None:
                    tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value = 'NULL'
        tbinfo.save("data/table_information.xlsx")
        print("table_information中添加该表.")
    else:
        print("table_information中没有该数据库.")
    
    #dbname.xlsx文件中增加一列
    if os.path.exists(db_path+using_dbname+'.xlsx'):
        db = load_workbook(db_path+using_dbname+'.xlsx')
        #tbname是否存在于sheet中
        if tbname in db.sheetnames:
            worksheet = db[tbname]
            #该表的最大列数
            max_column = worksheet.max_column
            #设置最大列+1，第一行的值为字段名
            column = columns_list[i].split(' ')
            worksheet.cell(row = 1,column= max_column+1).value = column[0]
            db.save(db_path+using_dbname+'.xlsx')
            print("已在数据库中添加该字段.")
        else:
            print("该表不在数据库中.")
    else:
        print("未找到数据库.")

#删除字段
def drop_field(tbname,columns_list,using_dbname,using_db):
    #查看是否存在该表
    #打开table_information
    db = load_workbook("data/table_information.xlsx")
    #是否存在该表
    if using_dbname in db.sheetnames:
        worksheet = db[using_dbname]

        length = len(columns_list)
        #删除匹配的行
        tbinfo = db
        tbinfo_tb = worksheet
        tbinfo_rows = tbinfo_tb.max_row
        for i in range(length):
            #找到匹配的行
            for j in range(tbinfo_rows):
                if tbinfo_tb.cell(row=j+1,column=1).value == tbname and tbinfo_tb.cell(row=j+1,column=2).value == columns_list[i]:
                    tbinfo_tb.delete_rows(j+1,1)
                    tbinfo.save("data/table_information.xlsx")
                    print("已在table_information中删除该字段.")
                elif j == tbinfo_rows-1 and tbinfo_tb.cell(row=j+1,column=1).value != tbname:
                    print("该表不在table_information中.")
    else:
        print("table_information中没有该数据库.")
    
    #dbname.xlsx文件中删除一列
    if os.path.exists(db_path+using_dbname+'.xlsx'):
        db = load_workbook(db_path+using_dbname+'.xlsx')
        #tbname是否存在于sheet中
        if tbname in db.sheetnames:
            worksheet = db[tbname]
            #查找匹配的第一个列元素
            length = len(columns_list)
            for i in range(length):
                columns = columns_list[i].split(' ')
                for j in range(worksheet.max_column):
                    if worksheet.cell(row=1,column=j+1).value == columns[0]:
                        worksheet.delete_cols(j+1,1)
                        db.save(db_path+using_dbname+'.xlsx')
                        print("已在数据库中删除该字段.")
                    elif j == worksheet.max_column-1 and worksheet.cell(row=1,column=j+1).value != columns[0]:
                        print("该表不在数据库中.")
        else:
            print("该表不在数据库中.")
    else:
        print("未找到数据库.")

#字段修改
def modify_field(tbname,alterFieldName,columns_list,using_dbname,using_db):
    #查看是否存在该表
    #打开table_information
    db = load_workbook("data/table_information.xlsx")
    #设置一个布尔变量
    flag = False

    #是否存在该表
    if using_dbname in db.sheetnames:
        worksheet = db[using_dbname]

        length = len(columns_list)
        #删除匹配的行
        tbinfo = db
        tbinfo_tb = worksheet
        tbinfo_rows = tbinfo_tb.max_row
        for i in range(length):             #将字段的属性写到table_information库中
            column = columns_list[i].split(' ')
            #找到匹配的行
            for j in range(tbinfo_rows):
                #检测是否已完成
                if flag == True:
                    break
                if tbinfo_tb.cell(row=j+1,column=1).value == tbname and tbinfo_tb.cell(row=j+1,column=2).value == alterFieldName:
                    tbinfo_tb.cell(row=j+1,column=2).value = column[0]
                    tbinfo_tb.cell(row=j+1,column=3).value = column[1]
                
                    for key in column[2:]:
                        if key == 'null':
                            tbinfo_tb.cell(row=j + 1, column=4).value = '1'
                        elif key == 'not_null':
                            tbinfo_tb.cell(row=j + 1, column=4).value = '0'
                        elif key == 'unique':
                            tbinfo_tb.cell(row=j + 1, column=5).value = '1'
                        elif key == 'pk':
                            tbinfo_tb.cell(row=j + 1, column=6).value = '1'
                        elif key == 'fk':
                            tbinfo_tb.cell(row=j + 1, column=7).value = '1'
            
                    for k in range(1, 8):
                        if tbinfo_tb.cell(row=j + 1, column=k).value is None:
                            tbinfo_tb.cell(row=j + 1, column=k).value = 'NULL'
                    flag = True
                #找到最后一行没有找到
                elif j == tbinfo_rows-1 and tbinfo_tb.cell(row=j+1,column=1).value != tbname:
                    print("该表不在数据库中.")
                elif j == tbinfo_rows-1 and tbinfo_tb.cell(row=j+1,column=1).value == tbname and tbinfo_tb.cell(row=j+1,column=2).value != column[0]:
                    print("该表中不存在该字段.")

        tbinfo.save("data/table_information.xlsx")
        print("已在table_information中修改该字段.")
    else:
        print("table_information中没有该数据库.")
    
    #dbname.xlsx文件中修改一列列名
    flag = False
    if os.path.exists(db_path+using_dbname+'.xlsx'):
        db = load_workbook(db_path+using_dbname+'.xlsx')
        #tbname是否存在于sheet中
        if tbname in db.sheetnames:
            worksheet = db[tbname]
            #查找匹配的第一个列元素
            length = len(columns_list)
            for i in range(length):
                columns = columns_list[i].split(' ')
                for j in range(worksheet.max_column):
                    if flag == True:
                        break
                    if worksheet.cell(row=1,column=j+1).value == alterFieldName:
                        worksheet.cell(row=1,column=j+1).value = columns[0]
                        db.save(db_path+using_dbname+'.xlsx')
                        print("已在数据库中修改该字段.")
                        flag = True
                    elif j == worksheet.max_column-1 and worksheet.cell(row=1,column=j+1).value != alterFieldName:
                        print("该表中不存在该字段.")
        else:
            print("该表不在数据库中.")
    else:
        print("未找到数据库.")

#插入
def insert_record(table_name, current_database, current_dbname, columns_list, multiFlag):
    #if not check_Constraint(columns,tablename,using_dbname,using_db):    #columns应为[dict]
        #print ("Constraint Error")
        #return False
    #找到标识
    flag = False
    #检查表名是否存在
    if multiFlag:
        if table_name in current_database.sheetnames:
            table = current_database[table_name]
            #columns_list本身为二维数组
            #columns为一维数组
            
            table_columns = table.max_column
            for columns in columns_list:
                table_rows = table.max_row
                for column in columns:
                    #查找匹配的列头是否与columns[0]匹配
                    for i in range(table_columns):
                        if i == 0:
                            flag = False
                        if flag == True:
                            break
                        if table.cell(row=1,column=i+1).value == column[0]:
                            #在最后一行插入一行值为columns[1]的数据
                            table.cell(row=table_rows+1,column=i+1).value = column[1]
                            #成功插入一行
                            print (column[0]+':'+column[1]+"插入成功.")
                            flag = True
                        elif i == table_columns-1 and table.cell(row=1,column=i+1).value != column[0]:
                            #没有找到对应的列头
                            print("该表中不存在该字段.")
        else:
            print("该表不在数据库中.")
    else:
        if table_name in current_database.sheetnames:
            table = current_database[table_name]
            #columns_list本身为二维数组
            #columns为一维数组
            table_rows = table.max_row
            table_columns = table.max_column
            for columns in columns_list:
                #查找匹配的列头是否与columns[0]匹配
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == columns[0]:
                        #在最后一行插入一行值为columns[1]的数据
                        table.cell(row=table_rows+1,column=i+1).value = columns[1]
                        #成功插入一行
                        print (columns[0]+':'+columns[1]+"插入成功.")
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != columns[0]:
                        #没有找到对应的列头
                        print("该表中不存在该字段.")
        else:
            print("该表不在数据库中.")
    current_database.save(db_path+current_dbname+'.xlsx')

#删除记录
def delete_record(table_name, current_database, current_dbname, condition_list):
    #检查表名是否存在
    if table_name in current_database.sheetnames:
        table = current_database[table_name]
        table_rows = table.max_row
        table_columns = table.max_column
        #二维数组
        delete_rows = []
        #一维数组
        delete_rows_list = []
        #查找与condition_list相符的行删除
        j = 0
        for condition in condition_list:
            #等于判断
            if '=' in condition:
                field_column = 0
                condition = condition.split('=')
                #找到列头=condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了condition[1]的值，记录在delete_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value == condition[1]:
                        delete_rows_list.append(i+1)
            #大于判断
            elif '>' in condition:
                field_column = 0
                condition = condition.split('>')
                #找到列头>condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了大于condition[1]的值，记录在delete_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value > condition[1]:
                        delete_rows_list.append(i+1)
            #小于判断
            elif '<' in condition:
                field_column = 0
                condition = condition.split('<')
                #找到列头<condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了小于condition[1]的值，记录在delete_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value < condition[1]:
                        delete_rows_list.append(i+1)
            delete_rows.append(delete_rows_list)
            j += 1

        
        #delete_rows没有元素
        if len(delete_rows) == 0:
            print("没有找到符合条件的记录.")
            return
        #将若有元素在delete_rows中每个组都出现，得出新的list
        for i in range(len(delete_rows)):
            for j in range(len(delete_rows[0])):
                flag = False
                for k in range(len(delete_rows[i])):
                    if delete_rows[0][j] == delete_rows[i][k]:
                        flag = False
                        break
                    else:
                        flag = True
            if flag:
                delete_rows[0].remove(delete_rows[0][j])
        #按照delete_rows[0]删除行
        for i in range(len(delete_rows[0])):
            table.delete_rows(delete_rows[0][i])
            print("第"+str(delete_rows[0][i]-1)+"行删除成功.")
        #保存xlsx
        current_database.save(db_path+current_dbname+'.xlsx')
    else:
        print("该表不在数据库中.")

#修改记录
def update_record(table_name,current_database,current_dbname,cols,condition_list,multiFlag):
    #检查表名是否存在
    if table_name in current_database.sheetnames:
        table = current_database[table_name]
        #查找符合condition_list的行进行修改
        table_rows = table.max_row
        table_columns = table.max_column
        #二维数组
        update_rows = []
        #一维数组
        update_rows_list = []
        #查找与condition_list相符的行修改
        j = 0
        for condition in condition_list:
            #等于判断
            if '=' in condition:
                field_column = 0
                condition = condition.split('=')
                #找到列头=condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了condition[1]的值，记录在update_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value == condition[1]:
                        update_rows_list.append(i+1)
            #大于判断
            elif '>' in condition:
                field_column = 0
                condition = condition.split('>')
                #找到列头>condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了大于condition[1]的值，记录在update_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value > condition[1]:
                        update_rows_list.append(i+1)
            #小于判断
            elif '<' in condition:
                field_column = 0
                condition = condition.split('<')
                #找到列头<condition[0]的列号
                for i in range(table_columns):
                    if i == 0:
                        flag = False
                    if flag == True:
                        break
                    if table.cell(row=1,column=i+1).value == condition[0]:
                        field_column = i+1
                        flag = True
                    elif i == table_columns-1 and table.cell(row=1,column=i+1).value != condition[0]:
                        print("该表中不存在该字段.")
                        return
                #若在第field_column列找到了小于condition[1]的值，记录在update_rows[j]中
                for i in range(table_rows):
                    if table.cell(row=i+1,column=field_column).value < condition[1]:
                        update_rows_list.append(i+1)
            update_rows.append(update_rows_list)
            j += 1
        #update_rows没有元素
        if len(update_rows) == 0:
            print("没有找到符合条件的记录.")
            return

        #将若有元素在update_rows中每个组都出现，得出新的list
        for i in range(len(update_rows)):
            for j in range(len(update_rows[0])):
                flag = False
                for k in range(len(update_rows[i])):
                    if update_rows[0][j] == update_rows[i][k]:
                        flag = False
                        break
                    else:
                        flag = True
            if flag:
                update_rows[0].remove(update_rows[0][j])

        #按照update_rows[0]修改行
        for i in range(len(update_rows[0])):
            if table_name in current_database.sheetnames:
                table = current_database[table_name]
                #columns_list本身为二维数组
                #columns为一维数组
                for columns in cols:
                    #查找匹配的列头是否与columns[0]匹配
                    for i in range(table_columns):
                        if i == 0:
                            flag = False
                        if flag == True:
                            break
                        if table.cell(row=1,column=i+1).value == columns[0]:
                            #在指定行插入一行值为columns[1]的数据
                            table.cell(row=update_rows[0][i],column=i+1).value = columns[1]
                            #成功插入一行
                            print (columns[0]+':'+columns[1]+"插入成功.")
                            flag = True
                        elif i == table_columns-1 and table.cell(row=1,column=i+1).value != columns[0]:
                            #没有找到对应的列头
                            print("该表中不存在该字段.")
                #保存xlsx文件
                current_database.save(db_path+current_dbname+'.xlsx')
            else:
                print("该表不在数据库中.")

#select a,b from table where c=x,d=x
def select(columns,table_name,using_dbname,using_db,limit={},predicate='and', symbol='=', tag=''):    #{'c':'x','d':'x'}
    if using_dbname == '':
        print ("please choose databse!")
        return
    table = using_db[table_name]
    #print columns
    if columns == '*' and len(limit) == 0:
        columns_name = list(iter_rows(table))[0]
        table_print = PrettyTable(columns_name)
        for i in range(1,len(list(iter_rows(table)))):
            table_print.add_row(list(iter_rows(table))[i])
        table_print.reversesort = True
        if tag == 'view':
            print (table_print)
            return list(iter_rows(table))   #view
        if tag == 'insert':
            return list(iter_rows(table))
        else:
            print(table_print)
    else:
        sel_cols = columns.split(',')   #*的情况
        rows_list = list(iter_rows(table))  #所有的行
        cols = rows_list[0]
        col_pos = []
        limit_pos = []
        print_row = []
        limit_cols = list(limit)
        symbol = '==' if symbol == '=' else symbol
        if columns[0] != '*':
            for i in range(len(sel_cols)):
                col_pos.append(cols.index(sel_cols[i])) #要查的列的列号
        else:
            sel_cols = list(iter_rows(table))[0]
            col_pos = range(len(cols))
        for i in range(len(limit)):
            limit_pos.append(cols.index(limit_cols[i])) #where的列
        for i in range(1, len(rows_list)):
            match = 0
            if predicate == 'in':
                match_list = limit[limit_cols[0]]
                for j in len(match_list):
                    if rows_list[i][limit_pos[0]] == match_list[j]:
                        print_row.append(i)
            if predicate == 'like':
                like_word = re.findall('(.*)\%',limit[limit_cols[0]])
                if like_word in rows_list[i][limit_pos[0]]:
                    print_row.append(i)
            else:
                for j in range(len(limit_pos)): #通过eval实现比较运算
                    if eval("'" + rows_list[i][limit_pos[j]] + "'" + symbol + "'" + limit[limit_cols[j]] + "'"):
                        match += 1
                if predicate == None:
                    print_row.append(i)
                if predicate == 'and' and match == len(limit_pos):  #and时要全部匹配
                    print_row.append(i)     #符合条件的行号
                if predicate == 'or' and match > 0: #or时至少一个匹配
                    print_row.append(i)

        table_print = PrettyTable(sel_cols)
        for i in range(len(print_row)):
            add_rows = []
            for x in col_pos:
                add_rows.append(rows_list[print_row[i]][x])
            table_print.add_row(add_rows)
        table_print.reversesort = True
        if tag == 'view':
            return table_print
        elif tag == 'insert':
            return table_print
        elif tag == 'nesting':
            tmpdb = using_db
            table = tmpdb['tmp']
            for i in range(len(sel_cols)):
                table.cell(row=0,column=i+1).value = sel_cols[i]
            for i in range(len(print_row)):
                add_rows = []
                for x in col_pos:
                    add_rows.append(rows_list[print_row[i]][x])
                for j in range(len(add_rows)):
                    table.cell(row=i+2,column=j+1).value = add_rows[j]
            tmpdb.save("data/" + using_dbname + ".xlsx")

        else:
            #table_print.reversesort = True
            print(table_print)

#grant select on test_tb for testuser
def set_permission(user,database,action):
    db = load_workbook("data/system.xlsx")
    table = db['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database) + 2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action) + 1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        print ("user have this permission")
    else:
        table.cell(row=row, column=col).value = table.cell(row=row, column=col).value + ',' + user
        db.save("data/system.xlsx")

#revoke select on test_tb for testuser
def del_permission(user,database,action):
    db = load_workbook("data/system.xlsx")
    table = db['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database) + 2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action) + 1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        if allow_user.index(user) == 0:
            table.cell(row=row, column=col).value = table.cell(row=row, column=col).value.replace(user, '')
        else:
            table.cell(row=row, column=col).value = table.cell(row=row, column=col).value.replace(',' + user, '')
        db.save("data/system.xlsx")
    else:
        print ("user didn't have this permission")

def check_permission(user,database,action):
    table = load_workbook("data/system.xlsx")['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database)+2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action)+1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        return True
    else:
        print ("Permission not allowed")
        return False

def check_syntax(sql):
    sql_words = sql.split(' ')
    for i in range(len(sql_words)):
        if sql_words[i] == 'select':
            if sql_words[i+2] == 'from':
                return True
        if sql_words[i] == 'from':
            if sql_words[i+2] == 'where':
                return True
def signup():
    return

def login(user,username,password,flagFirst,flagLogin):
    if check_login(username,password):
        print ("Login Success!Welcome {}! 😊".format(username))
        user = username
        flagLogin = True
        return user,flagFirst,flagLogin
    else:
        flagFirst = True
        print ("user not exist or password is wrong!😣 Try again.")
        login()
        return user,flagFirst,flagLogin

def check_login(username,password):
    db = load_workbook("data/system.xlsx")
    #right_pswd = select(password,user,{'username':username})
    table = db['user']
    col_list = list(iter_cols(table))
    try:
        pos = col_list[0].index(username)
    except:
        return False
    right_pswd = col_list[1][pos]
    if hashlib.md5(password.encode("utf-8")).hexdigest() == right_pswd:
        return True
    else:
        return False

def check_Constraint(columns,tablename,using_dbname,using_db):    #columns={'a':'xx'}
    db = load_workbook("data/table_information.xlsx")
    table = db[using_dbname]
    rows = []
    rows_list = list(iter_rows(table))  #所有行
    cols_list = list(iter_cols(table))
    for col in columns:
        value = col
        for i in range(len(cols_list[0])):  #table对应的行
            if cols_list[0][i] == tablename:
                rows.append(i)
        for line in rows:
            if rows_list[line][1] == col:
                typee, is_null, unique, pk, fk = rows_list[line][2:]
                if is_null == '0':
                    if value == '' or value.count(' ')>3:
                        return False
                if unique == '1':
                    if not check_unique(tablename,col,value,using_db):
                        return False
                if pk == '1':
                    if not check_unique(tablename,col,value,using_db) or value == '':
                        return False
                if '[' in typee:
                    typee, maxlen = re.findall(r'(\w*)\[(\d*)\]', typee) #int[10] => int,10
                else:
                    maxlen = 1000
                if len(value) > maxlen:
                    return False
                if typee == 'int':
                    if type(value) != type(1):
                        return False
                if typee == 'char':
                    if type(value) != type('c'):
                        return False
    return True

def check_unique(tablename,column,value,using_db):
    table = using_db[tablename]
    col_pos = list(iter_rows(table))[0].index(column)   #第几列
    cols_list = list(iter_cols(table))[col_pos][1:]
    if cols_list.count(value) > 1:  #该列中该值数量
        return False
    else:
        return True


def logout():
    return

def iter_rows(ws):      #表格按行数组形式输出，eg:list(iter_rows(a))
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def iter_cols(ws):      #表格按行数组形式输出，eg:list(iter_rows(a))
    for row in ws.iter_cols():
        yield [cell.value for cell in row]